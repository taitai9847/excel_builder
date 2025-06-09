import os
import csv
import config
import logging
from collections import defaultdict
from openpyxl import load_workbook

logger = logging.getLogger("excel_builder")

def count_columns(filename):
    with open(filename, encoding='utf-8') as f:
        reader = csv.reader(f)
        return len(next(reader))

def group_rows_by_file(input_csv, template_csv):
    grouped = defaultdict(list)
    input_col_count = count_columns(input_csv)
    template_col_count = count_columns(template_csv)
    with open(input_csv, encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if input_col_count != template_col_count:
                message = f"{input_csv} と {template_csv} のカラム数が一致していません: {row} 行目"
                logger.error(message)
                raise ValueError(message)
            file_name = row[0]
            grouped[file_name].append(row)
    return grouped

def write_values_from_input(input_csv=config.INPUT_FILE, template_csv=config.INPUT_TEMPLATE_FILE):
    input_col_count = count_columns(input_csv)
    template_col_count = count_columns(template_csv)
    if input_col_count != template_col_count:
        message = f"{input_csv} と {template_csv} のカラム数が一致していません"
        logger.error(message)
        raise ValueError(message)
    
    grouped_rows = group_rows_by_file(input_csv, template_csv)
    for file_name, rows in grouped_rows.items():
        output_path = "output/" + file_name
        if not os.path.exists(output_path):
            logger.error(f"Output フォルダにファイルが存在しません: {file_name}")
            continue
        try:
            output_path = "output/" + file_name
            wb = load_workbook(output_path)
            for row in rows:
                _, sheet_name, cell, value = row
                if sheet_name not in wb.sheetnames:
                    logger.error(f"シートが存在していません: {sheet_name} in {file_name}")
                    continue
                ws = wb[sheet_name]
                ws[cell] = str(value)
                logger.info(f"{file_name} - {sheet_name} の {cell} に '{value}' を入力しました")

            wb.save(output_path)
            wb.close()
            logger.info(f"{file_name} を保存しました")
        except Exception as e:
            logger.error(f"[{file_name}]:処理中にエラーが発生しました:{e}")
            continue
    return